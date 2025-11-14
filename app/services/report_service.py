"""
Report Generation Service

Generate and export reports from external systems
"""
from typing import Dict, Any, List, Optional
from loguru import logger
from datetime import datetime
import io


class ReportService:
    """
    Service for generating and exporting reports

    Features:
    - Generate reports from external systems
    - Export to PDF, Excel, CSV
    - Custom report templates
    - Scheduled reports
    """

    def __init__(self):
        self.report_formats = ['pdf', 'xlsx', 'csv', 'json']

    async def generate_report(
        self,
        adapter: Any,
        report_name: str,
        model: str,
        record_ids: List[int],
        format: str = 'pdf'
    ) -> bytes:
        """
        Generate report from external system

        Args:
            adapter: System adapter
            report_name: Report template name
            model: Model name
            record_ids: List of record IDs
            format: Export format (pdf, xlsx, csv)

        Returns:
            Report content as bytes

        Example:
            report_content = await report_service.generate_report(
                adapter=odoo_adapter,
                report_name="sale.report_saleorder",
                model="sale.order",
                record_ids=[1, 2, 3],
                format="pdf"
            )
        """
        try:
            if format not in self.report_formats:
                raise ValueError(f"Unsupported format: {format}. Allowed: {self.report_formats}")

            # For Odoo systems
            if hasattr(adapter, 'call_method'):
                # Generate report using Odoo's report system
                result = await adapter.call_method(
                    model="ir.actions.report",
                    method="_render_qweb_pdf",
                    args=[report_name, record_ids]
                )

                if isinstance(result, (list, tuple)) and len(result) > 0:
                    report_content = result[0]
                else:
                    report_content = result

                logger.info(f"Generated report {report_name} for {len(record_ids)} records")

                return report_content

            else:
                raise NotImplementedError("Report generation not supported for this system type")

        except Exception as e:
            logger.error(f"Report generation error: {str(e)}")
            raise

    async def export_to_excel(
        self,
        data: List[Dict[str, Any]],
        filename: str = "export.xlsx"
    ) -> bytes:
        """
        Export data to Excel format

        Args:
            data: List of records to export
            filename: Output filename

        Returns:
            Excel file content as bytes

        Example:
            excel_content = await report_service.export_to_excel(
                data=[
                    {"name": "Ahmed", "email": "ahmed@example.com"},
                    {"name": "Sara", "email": "sara@example.com"}
                ],
                filename="partners.xlsx"
            )
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment

            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Export"

            if not data:
                return b""

            # Write headers
            headers = list(data[0].keys())
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Write data
            for row_num, record in enumerate(data, 2):
                for col_num, header in enumerate(headers, 1):
                    value = record.get(header, "")
                    # Handle datetime objects
                    if isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d %H:%M:%S")
                    ws.cell(row=row_num, column=col_num, value=value)

            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            logger.info(f"Exported {len(data)} records to Excel")

            return output.getvalue()

        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            raise
        except Exception as e:
            logger.error(f"Excel export error: {str(e)}")
            raise

    async def export_to_csv(
        self,
        data: List[Dict[str, Any]],
        filename: str = "export.csv"
    ) -> bytes:
        """
        Export data to CSV format

        Args:
            data: List of records to export
            filename: Output filename

        Returns:
            CSV file content as bytes
        """
        try:
            import csv

            if not data:
                return b""

            output = io.StringIO()
            headers = list(data[0].keys())

            writer = csv.DictWriter(output, fieldnames=headers)
            writer.writeheader()

            for record in data:
                # Handle datetime objects
                formatted_record = {}
                for key, value in record.items():
                    if isinstance(value, datetime):
                        formatted_record[key] = value.strftime("%Y-%m-%d %H:%M:%S")
                    else:
                        formatted_record[key] = value
                writer.writerow(formatted_record)

            logger.info(f"Exported {len(data)} records to CSV")

            return output.getvalue().encode('utf-8')

        except Exception as e:
            logger.error(f"CSV export error: {str(e)}")
            raise

    async def generate_sales_report(
        self,
        adapter: Any,
        start_date: datetime,
        end_date: datetime,
        format: str = 'pdf'
    ) -> bytes:
        """
        Generate sales report

        Args:
            adapter: System adapter
            start_date: Report start date
            end_date: Report end date
            format: Export format

        Returns:
            Report content
        """
        try:
            # Search for sale orders in date range
            domain = [
                ["date_order", ">=", start_date.strftime("%Y-%m-%d")],
                ["date_order", "<=", end_date.strftime("%Y-%m-%d")],
                ["state", "in", ["sale", "done"]]
            ]

            orders = await adapter.search_read(
                model="sale.order",
                domain=domain,
                fields=["name", "partner_id", "date_order", "amount_total", "state"]
            )

            if format == 'xlsx':
                return await self.export_to_excel(orders, "sales_report.xlsx")
            elif format == 'csv':
                return await self.export_to_csv(orders, "sales_report.csv")
            elif format == 'pdf':
                # Generate PDF report
                record_ids = [order['id'] for order in orders]
                return await self.generate_report(
                    adapter=adapter,
                    report_name="sale.report_saleorder",
                    model="sale.order",
                    record_ids=record_ids,
                    format="pdf"
                )

        except Exception as e:
            logger.error(f"Sales report error: {str(e)}")
            raise

    async def generate_inventory_report(
        self,
        adapter: Any,
        product_ids: Optional[List[int]] = None,
        format: str = 'xlsx'
    ) -> bytes:
        """
        Generate inventory report

        Args:
            adapter: System adapter
            product_ids: Optional list of product IDs
            format: Export format

        Returns:
            Report content
        """
        try:
            domain = []
            if product_ids:
                domain.append(["id", "in", product_ids])

            products = await adapter.search_read(
                model="product.product",
                domain=domain,
                fields=[
                    "name", "default_code", "barcode",
                    "qty_available", "virtual_available",
                    "list_price", "standard_price"
                ]
            )

            if format == 'xlsx':
                return await self.export_to_excel(products, "inventory_report.xlsx")
            elif format == 'csv':
                return await self.export_to_csv(products, "inventory_report.csv")

        except Exception as e:
            logger.error(f"Inventory report error: {str(e)}")
            raise

    async def generate_partner_report(
        self,
        adapter: Any,
        is_company: Optional[bool] = None,
        format: str = 'xlsx'
    ) -> bytes:
        """
        Generate partner/customer report

        Args:
            adapter: System adapter
            is_company: Filter by company/individual
            format: Export format

        Returns:
            Report content
        """
        try:
            domain = []
            if is_company is not None:
                domain.append(["is_company", "=", is_company])

            partners = await adapter.search_read(
                model="res.partner",
                domain=domain,
                fields=[
                    "name", "email", "phone", "mobile",
                    "street", "city", "country_id",
                    "vat", "is_company"
                ]
            )

            if format == 'xlsx':
                return await self.export_to_excel(partners, "partners_report.xlsx")
            elif format == 'csv':
                return await self.export_to_csv(partners, "partners_report.csv")

        except Exception as e:
            logger.error(f"Partner report error: {str(e)}")
            raise
